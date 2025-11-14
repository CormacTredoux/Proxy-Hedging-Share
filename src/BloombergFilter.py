import pandas as pd
import openpyxl
import re
import os

def clean_bloomberg_excel(file_path): ### note "volume" not used anymore
    # Load Excel file
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    nrows = sheet.max_row

    start_idx = None
    for i in range(1, nrows + 1):
        first_cell = sheet.cell(row=i, column=1).value
        if first_cell is None:
            continue
        first_cell_str = str(first_cell).strip().lower()
        if first_cell_str == "date":
            start_idx = i
            break

    if start_idx is None:
        raise ValueError(f"No 'Date' header found in {file_path}")

    df = pd.read_excel(file_path, skiprows=start_idx - 1, header=0)
    df.replace(",", ".", regex=True, inplace=True)

    # Check for required columns
    required_cols = ['Date', 'PX_LAST']
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"'{col}' column not found in {file_path}")

    # Check for 'Volume' and include if present
    include_volume = 'Volume' in df.columns
    cols_to_keep = ['Date', 'PX_LAST']
    if include_volume:
        cols_to_keep.append('Volume')
    
    df = df[cols_to_keep]
    df['PX_LAST'] = df['PX_LAST'].astype(float)
    df['Date'] = pd.to_datetime(df['Date'])
    
    # Convert 'Volume' to float if it exists
    if include_volume:
        df['Volume'] = pd.to_numeric(df['Volume'], errors='coerce')

    # Drop any rows with a missing value in any column
    df.dropna(how="any", inplace=True)

    # Drop the most recent price, which is often bugged
    df = df.iloc[1:].reset_index(drop=True)

    return df, include_volume


### Below is the script to clean all (EXCEL) files in a folder
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)

raw_dir = os.path.join(project_root, "Raw_Dat")
cleaned_dir = os.path.join(project_root, "Clean_Dat")
os.makedirs(cleaned_dir, exist_ok=True)
cleaned_data = {}
has_volume_data = {}

# Process each .xlsx file
for filename in os.listdir(raw_dir):
    if filename.endswith(".xlsx"):
        full_path = os.path.join(raw_dir, filename)
        try:
            df, has_volume = clean_bloomberg_excel(full_path)
            cleaned_data[filename] = df
            has_volume_data[filename] = has_volume

            # Save cleaned data
            target_path = os.path.join(cleaned_dir, filename)
            df.to_excel(target_path, index=False)

            print(f"Successful cleaning: {filename}")
        except Exception as e:
            print(f"Failed to clean {filename}: {e}")

# Combine PX_LAST data
combined_pxlast_df = pd.DataFrame()
for filename, df in cleaned_data.items():
    temp_df = df[['Date', 'PX_LAST']].copy()
    equity_name = filename.replace('.xlsx', '')
    temp_df.rename(columns={'PX_LAST': equity_name}, inplace=True)
    
    if combined_pxlast_df.empty:
        combined_pxlast_df = temp_df
    else:
        combined_pxlast_df = pd.merge(combined_pxlast_df, temp_df, on='Date', how='outer')

combined_pxlast_df.dropna(how="any", inplace=True)
combined_pxlast_file_path = os.path.join(cleaned_dir, 'Combined_data.xlsx')
combined_pxlast_df.to_excel(combined_pxlast_file_path, index=False)
print(f"Combined PX_LAST data saved as: {combined_pxlast_file_path}")

# Combine Volume data
combined_volume_df = pd.DataFrame()
for filename, df in cleaned_data.items():
    if has_volume_data[filename]:
        temp_df = df[['Date', 'Volume']].copy()
        equity_name = filename.replace('.xlsx', '')
        temp_df.rename(columns={'Volume': equity_name}, inplace=True)
        
        if combined_volume_df.empty:
            combined_volume_df = temp_df
        else:
            combined_volume_df = pd.merge(combined_volume_df, temp_df, on='Date', how='outer')

if not combined_volume_df.empty:
    combined_volume_file_path = os.path.join(cleaned_dir, 'Combined_data_vol.xlsx')
    combined_volume_df.to_excel(combined_volume_file_path, index=False)
    print(f"Combined Volume data saved as: {combined_volume_file_path}")
else:
    print("No files with a 'Volume' column found. 'Combined_data_vol.xlsx' was not created.")